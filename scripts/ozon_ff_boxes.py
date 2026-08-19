#!/usr/bin/env python3
"""Выгрузка «какой товар в какой короб» для фулфилмента из Ozon Seller API → Excel."""

from __future__ import annotations

import argparse
import json
import os
import time
from datetime import date, datetime
from pathlib import Path

import requests
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
EXPORTS = Path(os.environ.get("OZON_FF_EXPORTS_DIR", str(ROOT / "Выгрузки")))
DASHBOARD = Path(os.environ.get("OZON_FF_DASHBOARD_DIR", str(ROOT / "Дашборд_FF")))
ENV_FILE = Path(os.environ.get("OZON_FF_ENV_FILE", str(ROOT / ".env")))
API = "https://api-seller.ozon.ru"

STATE_LABELS = {
    "READY_TO_SUPPLY": "Готова к отгрузке",
    "DATA_FILLING": "Заполнение данных",
    "COMPLETED": "Завершена",
    "IN_TRANSIT": "В пути",
}

ZONE_LABELS = {
    "SORT": "Сортируемый",
    "NON_SORT": "Несортируемый",
    "UNSPECIFIED": "—",
}


def load_env() -> tuple[str, str]:
    if ENV_FILE.exists():
        for line in ENV_FILE.read_text(encoding="utf-8").splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            os.environ.setdefault(k.strip(), v.strip())
    cid = os.environ.get("OZON_CLIENT_ID", "")
    key = os.environ.get("OZON_API_KEY", "")
    if not cid or not key:
        raise SystemExit(
            f"Нужны OZON_CLIENT_ID и OZON_API_KEY в {ENV_FILE} "
            f"(см. {ENV_FILE.with_suffix('.example')})"
        )
    return cid, key


LIST_STATES = [
    "READY_TO_SUPPLY",
    "DATA_FILLING",
    "IN_TRANSIT",
    "ACCEPTANCE_AT_STORAGE_WAREHOUSE",
    "COMPLETED",
]


def slug_order_number(order_number: str) -> str:
    return str(order_number).replace("/", "-").strip()


class OzonClient:
    def __init__(self, client_id: str, api_key: str) -> None:
        self.headers = {
            "Client-Id": client_id,
            "Api-Key": api_key,
            "Content-Type": "application/json",
        }

    def post(self, path: str, body: dict, retries: int = 5) -> dict:
        delay = 0.4
        for attempt in range(retries):
            r = requests.post(f"{API}{path}", headers=self.headers, json=body, timeout=60)
            if r.status_code == 429:
                time.sleep(delay * (attempt + 2))
                continue
            r.raise_for_status()
            data = r.json()
            if isinstance(data, dict) and data.get("code") not in (None, 0):
                raise RuntimeError(f"{path}: {data.get('message', data)}")
            time.sleep(delay)
            return data
        r.raise_for_status()
        return {}

    def list_recent_order_ids(self, limit: int = 40, states: list[str] | None = None) -> list[int]:
        states = states or LIST_STATES
        ids: list[int] = []
        last_id = ""
        while len(ids) < limit:
            data = self.post(
                "/v3/supply-order/list",
                {
                    "filter": {"states": states},
                    "last_id": last_id,
                    "limit": min(50, limit - len(ids)),
                    "sort_by": "ORDER_CREATION",
                    "sort_direction": "DESC",
                },
            )
            batch = data.get("order_ids", [])
            if not batch:
                break
            ids.extend(batch)
            last_id = data.get("last_id", "")
            if not last_id:
                break
        return ids[:limit]

    def list_active_orders(self, states: list[str] | None = None) -> list[int]:
        return self.list_recent_order_ids(limit=50, states=states or ["READY_TO_SUPPLY", "DATA_FILLING"])

    def find_order_id_by_number(self, order_number: str, scan_limit: int = 200) -> int | None:
        target = slug_order_number(order_number)
        for order_id in self.list_recent_order_ids(limit=scan_limit):
            order = self.get_order(order_id)
            if slug_order_number(order.get("order_number", "")) == target:
                return order_id
        return None

    def get_order(self, order_id: int) -> dict:
        data = self.post("/v3/supply-order/get", {"order_ids": [order_id]})
        orders = data.get("orders", [])
        if not orders:
            raise RuntimeError(f"Поставка {order_id} не найдена")
        return orders[0]

    def get_cargoes(self, supply_id: int) -> dict:
        data = self.post("/v2/cargoes/get", {"supplies": [{"supply_id": supply_id}]})
        supplies = data.get("supplies", [])
        return supplies[0] if supplies else {}

    def get_bundle_items(self, bundle_id: str) -> list[dict]:
        items: list[dict] = []
        last_id = ""
        while True:
            data = self.post(
                "/v1/supply-order/bundle",
                {"bundle_ids": [bundle_id], "limit": 100, "last_id": last_id},
            )
            items.extend(data.get("items", []))
            if not data.get("has_next"):
                break
            last_id = data.get("last_id", "")
            if not last_id:
                break
        return items

    def get_cluster_names(self) -> dict[str, str]:
        """macrolocal_cluster_id → «Пермь», «Дальний Восток» и т.д."""
        names: dict[str, str] = {}
        for cluster_type in ("CLUSTER_TYPE_OZON", "CLUSTER_TYPE_CIS"):
            data = self.post("/v1/cluster/list", {"cluster_type": cluster_type})
            for cluster in data.get("clusters", []):
                macro_id = cluster.get("macrolocal_cluster_id")
                name = cluster.get("name", "")
                if macro_id is not None and name:
                    names[str(macro_id)] = name
        return names


def fmt_timeslot(order: dict) -> str:
    ts = (order.get("timeslot") or {}).get("timeslot") or {}
    frm, to = ts.get("from"), ts.get("to")
    if not frm:
        return "—"
    try:
        f = datetime.fromisoformat(frm.replace("Z", "+00:00"))
        t = datetime.fromisoformat(to.replace("Z", "+00:00")) if to else None
        if t:
            return f"{f.strftime('%d.%m.%Y %H:%M')} – {t.strftime('%H:%M')}"
        return f.strftime("%d.%m.%Y %H:%M")
    except ValueError:
        return f"{frm} – {to or ''}"


def cluster_label(cluster_id: str | int, names: dict[str, str]) -> str:
    key = str(cluster_id)
    return names.get(key, f"кластер {key}")


def collect_boxes(
    client: OzonClient, order: dict, cluster_names: dict[str, str]
) -> tuple[dict, list[dict], list[dict]]:
    meta = {
        "order_id": order.get("order_id"),
        "order_number": order.get("order_number"),
        "state": STATE_LABELS.get(order.get("state", ""), order.get("state", "")),
        "created": (order.get("created_date") or "")[:10],
        "timeslot": fmt_timeslot(order),
        "dropoff": (order.get("drop_off_warehouse") or {}).get("name", "—"),
        "dropoff_address": (order.get("drop_off_warehouse") or {}).get("address", ""),
    }

    by_box: list[dict] = []
    by_sku: list[dict] = []
    box_no = 0

    for supply in order.get("supplies", []):
        supply_id = supply.get("supply_id")
        cluster_id = supply.get("macrolocal_cluster_id", "")
        region = cluster_label(cluster_id, cluster_names)
        cargo_info = client.get_cargoes(supply_id)
        cargoes = cargo_info.get("cargoes", [])

        for cargo in cargoes:
            box_no += 1
            bundle_id = cargo.get("bundle_id", "")
            barcode = str(cargo.get("cargo_id", ""))
            cargo_type = cargo.get("type", "BOX")
            zone = ZONE_LABELS.get(
                (cargo.get("placement_zone_type") or "").replace("TYPE_", ""),
                cargo.get("placement_zone_type", ""),
            )

            items = client.get_bundle_items(bundle_id) if bundle_id else []
            if not items:
                by_box.append(
                    {
                        "Короб №": box_no,
                        "Штрихкод короба": barcode,
                        "Куда едет": region,
                        "Supply ID": supply_id,
                        "Артикул": "—",
                        "Название": "⚠️ пустой короб / нет состава",
                        "Шт": 0,
                        "Зона": zone,
                        "Тип": cargo_type,
                    }
                )
                continue

            for item in items:
                row = {
                    "Короб №": box_no,
                    "Штрихкод короба": barcode,
                    "Куда едет": region,
                    "Supply ID": supply_id,
                    "Артикул": item.get("offer_id", ""),
                    "Название": item.get("name", ""),
                    "Шт": item.get("quantity", 0),
                    "Зона": ZONE_LABELS.get(item.get("placement_zone", ""), item.get("placement_zone", "")),
                    "Тип": cargo_type,
                    "SKU Ozon": item.get("sku", ""),
                    "Штрихкод товара": item.get("barcode", ""),
                }
                by_box.append(row)
                by_sku.append(
                    {
                        "Артикул": item.get("offer_id", ""),
                        "Название": item.get("name", ""),
                        "Шт всего": item.get("quantity", 0),
                        "→ Короб №": box_no,
                        "Штрихкод короба": barcode,
                        "Куда едет": region,
                    }
                )

    by_sku.sort(key=lambda r: (str(r["Артикул"]).lower(), r["→ Короб №"]))
    return meta, by_box, by_sku


def style_header(ws, row: int = 1) -> None:
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def autosize(ws, max_width: int = 55) -> None:
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        width = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[letter].width = min(max(width + 2, 10), max_width)


def write_sheet_info(wb: Workbook, meta: dict, box_count: int, sku_rows: int) -> None:
    ws = wb.active
    ws.title = "Инфо"
    ws["A1"] = "Поставка Ozon — для фулфилмента"
    ws["A1"].font = Font(bold=True, size=14)
    rows = [
        ("Номер поставки", meta["order_number"]),
        ("ID заявки", meta["order_id"]),
        ("Статус", meta["state"]),
        ("Создана", meta["created"]),
        ("Слот отгрузки", meta["timeslot"]),
        ("Точка сдачи", meta["dropoff"]),
        ("Адрес", meta["dropoff_address"]),
        ("Коробов", box_count),
        ("Строк товаров", sku_rows),
        ("Выгрузка", date.today().isoformat()),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws[f"A{i}"] = k
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"] = v
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60


def write_table_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict]) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    style_header(ws)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    ws.freeze_panes = "A2"
    autosize(ws)


def build_excel(meta: dict, by_box: list[dict], by_sku: list[dict], output: Path) -> Path:
    output.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    box_numbers = {r["Короб №"] for r in by_box if r.get("Короб №")}
    write_sheet_info(wb, meta, len(box_numbers), len(by_sku))

    box_headers = [
        "Короб №",
        "Штрихкод короба",
        "Артикул",
        "Название",
        "Шт",
        "Куда едет",
        "Зона",
        "Supply ID",
    ]
    write_table_sheet(wb, "ПО КОРОБАМ", box_headers, by_box)

    sku_headers = ["Артикул", "Название", "Шт всего", "→ Короб №", "Штрихкод короба", "Куда едет"]
    write_table_sheet(wb, "ПО ТОВАРУ", sku_headers, by_sku)

    wb.save(output)
    return output


def build_web_payload(meta: dict, by_box: list[dict], by_sku: list[dict]) -> dict:
    boxes_map: dict[int, dict] = {}
    for row in by_box:
        no = row["Короб №"]
        if no not in boxes_map:
            boxes_map[no] = {
                "no": no,
                "barcode": row.get("Штрихкод короба", ""),
                "region": row.get("Куда едет", ""),
                "zone": row.get("Зона", ""),
                "items": [],
            }
        if row.get("Артикул") != "—":
            boxes_map[no]["items"].append(
                {
                    "article": row.get("Артикул", ""),
                    "name": row.get("Название", ""),
                    "qty": row.get("Шт", 0),
                }
            )
    return {
        "meta": {**meta, "generated": date.today().isoformat(), "box_count": len(boxes_map)},
        "boxes": [boxes_map[k] for k in sorted(boxes_map)],
        "by_sku": by_sku,
    }


def save_web_data(payload: dict, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return path


def build_web_site(client: OzonClient, dashboard_dir: Path = DASHBOARD, limit: int = 40) -> Path:
    """Каталог поставок + json/xlsx на каждую — для сайта с выбором по номеру."""
    supplies_dir = dashboard_dir / "supplies"
    supplies_dir.mkdir(parents=True, exist_ok=True)
    cluster_names = client.get_cluster_names()
    order_ids = client.list_recent_order_ids(limit=limit)

    catalog_entries: list[dict] = []
    for order_id in order_ids:
        try:
            order = client.get_order(order_id)
            meta, by_box, by_sku = collect_boxes(client, order, cluster_names)
        except (RuntimeError, requests.HTTPError) as exc:
            print(f"  skip {order_id}: {exc}")
            continue
        if not by_box:
            continue

        slug = slug_order_number(meta["order_number"])
        payload = build_web_payload(meta, by_box, by_sku)
        save_web_data(payload, supplies_dir / f"{slug}.json")
        build_excel(meta, by_box, by_sku, supplies_dir / f"{slug}.xlsx")

        catalog_entries.append(
            {
                "slug": slug,
                "order_number": meta["order_number"],
                "order_id": meta["order_id"],
                "state": meta["state"],
                "timeslot": meta["timeslot"],
                "box_count": payload["meta"]["box_count"],
                "created": meta["created"],
            }
        )
        print(f"  ✓ {meta['order_number']} — {payload['meta']['box_count']} коробов")

    if not catalog_entries:
        raise SystemExit("Нет поставок с коробами для публикации")

    catalog = {
        "generated": date.today().isoformat(),
        "supplies": catalog_entries,
        "default_slug": catalog_entries[0]["slug"],
    }
    catalog_path = dashboard_dir / "catalog.json"
    catalog_path.write_text(json.dumps(catalog, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"✓ каталог: {len(catalog_entries)} поставок → {catalog_path}")
    return catalog_path


def fetch_and_build(
    client: OzonClient, order_id: int | None = None, order_number: str | None = None
) -> tuple[dict, list[dict], list[dict], int]:
    if order_id is None and order_number:
        order_id = client.find_order_id_by_number(order_number)
        if order_id is None:
            raise SystemExit(f"Поставка {order_number} не найдена в последних заявках Ozon")
    if order_id is None:
        ids = client.list_active_orders()
        if not ids:
            raise SystemExit("Нет активных поставок (READY_TO_SUPPLY / DATA_FILLING)")
        order_id = ids[0]
        if len(ids) > 1:
            print(f"Найдено {len(ids)} поставок, берём первую: {order_id}")
    order = client.get_order(order_id)
    cluster_names = client.get_cluster_names()
    meta, by_box, by_sku = collect_boxes(client, order, cluster_names)
    if not by_box:
        raise SystemExit(f"В поставке {order_id} нет грузомест (коробов)")
    return meta, by_box, by_sku, order_id


def main() -> None:
    parser = argparse.ArgumentParser(description="Excel для ФФ: короб → товар (Ozon API)")
    parser.add_argument("--order-id", type=int, help="ID заявки на поставку (иначе — активная)")
    parser.add_argument("--order-number", help="Номер поставки, например 117111449-1")
    parser.add_argument("-o", "--output", type=Path, help="Путь к .xlsx")
    parser.add_argument("--build-web", action="store_true", help="Собрать каталог поставок для сайта")
    parser.add_argument("--web-limit", type=int, default=80, help="Сколько поставок сканировать для каталога")
    parser.add_argument("--web-only", action="store_true", help="Только web-данные, без Excel в Выгрузки")
    parser.add_argument("--no-web", action="store_true", help="Не писать web-данные")
    args = parser.parse_args()

    cid, key = load_env()
    client = OzonClient(cid, key)

    if args.build_web:
        build_web_site(client, limit=args.web_limit)
        return

    meta, by_box, by_sku, order_id = fetch_and_build(
        client, args.order_id, args.order_number
    )

    if not args.no_web:
        slug = slug_order_number(meta["order_number"])
        payload = build_web_payload(meta, by_box, by_sku)
        save_web_data(payload, DASHBOARD / "supplies" / f"{slug}.json")
        print(f"✓ {DASHBOARD / 'supplies' / slug}.json")

    if args.web_only:
        print(f"  Коробов: {len({r['Короб №'] for r in by_box})}, строк: {len(by_box)}")
        return

    order_num = slug_order_number(meta.get("order_number", str(order_id)))
    out = args.output or EXPORTS / f"FF_короба_Ozon_{order_num}_{date.today().isoformat()}.xlsx"
    path = build_excel(meta, by_box, by_sku, out)
    print(f"✓ {path}")
    print(f"  Коробов: {len({r['Короб №'] for r in by_box})}, строк: {len(by_box)}")


if __name__ == "__main__":
    main()
